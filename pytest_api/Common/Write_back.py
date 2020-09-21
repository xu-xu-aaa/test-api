from openpyxl import load_workbook
from Common.Reflection import get_data


class Write_Back:
    def set_data(self, path_name, sheet_variable):
        # 变量结果写回
        variable_data = getattr(get_data, "variable_data")
        wb = load_workbook(path_name)
        sheet_variable = wb[sheet_variable]
        h = 0
        for key in variable_data.keys():
            h = h + 1
            sheet_variable.cell(h, 1).value = key
            sheet_variable.cell(h, 2).value = variable_data[key]
        # 响应结果写回
        response_data = get_data.response
        for sheet_data in response_data:
            for sheet_key in sheet_data:
                sheet = wb[sheet_key]
                data_sheet = sheet_data[sheet_key]
                for case_id in data_sheet:
                    data = data_sheet[case_id]
                    print(data)
                    for j in range(0, len(data)):
                        if j == 0:
                            i = 0
                            for value_name in getattr(get_data, "title_case"):
                                i = i + 1
                                if str(value_name) == "result":
                                    sheet.cell(int(case_id), i).value = data[j]
                                    continue
                        elif j == 1:
                            i = 0
                            for value_name in getattr(get_data, "title_case"):
                                i = i + 1
                                if str(value_name) == "response":
                                    sheet.cell(int(case_id), i).value = str(data[j])
                                    continue
                        elif j == 2:
                            i = 0
                            for value_name in getattr(get_data, "title_case"):
                                i = i + 1
                                if str(value_name) == "sql_response":
                                    sheet.cell(int(case_id), i).value = str(data[j])
                                    continue
                        elif j == 3:
                            i = 0
                            for value_name in getattr(get_data, "title_case"):
                                i = i + 1
                                if str(value_name) == "sql_result":
                                    sheet.cell(int(case_id), i).value = data[j]
                                    continue
                        else:
                            raise
        wb.save(path_name)