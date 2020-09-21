from openpyxl import load_workbook
from Common.Reflection import get_data


class Case:
    def get_case(self, file_name, sheet_name):  # 读excel测试用例
        wb = load_workbook(file_name)
        title = []
        sheet = wb[sheet_name]
        for j in range(1, sheet.max_column + 1):
            title.append(sheet.cell(1, j).value)
        setattr(get_data, "title_case", title)
        test_case = []
        for i in range(3, sheet.max_row + 1):
            data = {}
            for j in range(1, sheet.max_column + 1):
                if title[j - 1] is not None:
                    data[title[j - 1]] = sheet.cell(i, j).value
            test_case.append(data)
        return test_case

    def get_variable(self, file_name, sheet_name):  # 读excel变量基础数据
        wb = load_workbook(file_name)
        title = []
        data = {}
        sheet = wb[sheet_name]
        for j in range(1, sheet.max_row + 1):
            title.append(sheet.cell(j, 1).value)
        for j in range(1, sheet.max_row + 1):
            if title[j - 1] is not None:
                data[title[j - 1]] = sheet.cell(j, 2).value
        return data